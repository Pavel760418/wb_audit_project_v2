"""
Низкоуровневые примитивы сборки итогового Excel-файла:
светофор, статусы блоков, условное форматирование, базовые таблицы.
"""
from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

STATUS_COLORS = {
    "ok": "C6EFCE",
    "full": "C6EFCE",
    "warning": "FFEB9C",
    "partial": "FFEB9C",
    "missing": "FFC7CE",
    "error": "FFC7CE",
    "none": "FFC7CE",
    "insufficient_data": "FFC7CE",
}

STATUS_LABELS = {
    "ok": "Полный расчёт",
    "full": "Полный расчёт",
    "warning": "Частичный расчёт",
    "partial": "Частичный расчёт",
    "missing": "Нет данных для расчёта",
    "error": "Нет данных для расчёта",
    "none": "Нет данных для расчёта",
    "insufficient_data": "Нет данных для расчёта",
}


def write_sheet_status(ws, status: str, cell: str = "A1"):
    label = STATUS_LABELS.get(status, "Нет данных для расчёта")
    color = STATUS_COLORS.get(status, "FFC7CE")
    ws[cell] = f"Статус листа: {label}"
    ws[cell].font = Font(bold=True, size=12)
    ws[cell].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")


def write_availability_table(ws, summary_rows: list, start_row: int = 3):
    headers = ["Отчёт", "Обязателен", "Статус", "Комментарий"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=col_idx, value=header)
        cell.font = Font(bold=True)

    for row_offset, row in enumerate(summary_rows, start=1):
        r = start_row + row_offset
        ws.cell(row=r, column=1, value=row["Отчёт"])
        ws.cell(row=r, column=2, value=row["Обязателен"])
        status_cell = ws.cell(row=r, column=3, value=STATUS_LABELS.get(row["Статус"], row["Статус"]))
        color = STATUS_COLORS.get(row["Статус"], "FFC7CE")
        status_cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        ws.cell(row=r, column=4, value=row["Комментарий"])

    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 32


def autosize_columns(ws, max_width: int = 60):
    for column_cells in ws.columns:
        length = max((len(str(cell.value)) if cell.value is not None else 0) for cell in column_cells)
        col_letter = get_column_letter(column_cells[0].column)
        ws.column_dimensions[col_letter].width = min(max_width, max(12, length + 2))


# Формат денежных значений в рублях (два знака после запятой, разделитель разрядов).
MONEY_FORMAT = "#,##0.00 ₽"
INT_FORMAT = "#,##0"


def write_title(ws, text: str, row: int, col: int = 1, size: int = 12, bold: bool = True) -> int:
    """Пишет заголовок и возвращает следующую свободную строку."""
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = Font(bold=bold, size=size)
    return row + 1


def write_note(ws, text: str, row: int, col: int = 1, italic: bool = True) -> int:
    """Пишет пояснительный текст (например, причину частичного расчёта)."""
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = Font(italic=italic)
    cell.alignment = Alignment(wrap_text=True, vertical="top")
    return row + 1


def write_kpi_table(ws, pairs: list, start_row: int, headers=("Показатель", "Значение")) -> int:
    """Пишет таблицу «Показатель / Значение» с денежным форматированием.

    pairs: список кортежей (label, value, kind), где kind ∈ {"money", "int", "text"}.
    Допускается и (label, value) — тогда тип определяется автоматически.
    Возвращает следующую свободную строку.
    """
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=col_idx, value=header)
        cell.font = Font(bold=True)

    r = start_row + 1
    for pair in pairs:
        if len(pair) == 3:
            label, value, kind = pair
        else:
            label, value = pair
            kind = "money" if isinstance(value, float) else ("int" if isinstance(value, int) else "text")

        ws.cell(row=r, column=1, value=label)
        value_cell = ws.cell(row=r, column=2, value=value)
        if kind == "money" and isinstance(value, (int, float)):
            value_cell.number_format = MONEY_FORMAT
        elif kind == "int" and isinstance(value, (int, float)):
            value_cell.number_format = INT_FORMAT
        r += 1

    return r


def write_simple_table(ws, headers: list, rows: list, start_row: int, money_columns=None) -> int:
    """Пишет обычную таблицу с жирной шапкой. rows — список списков значений.

    money_columns — множество индексов колонок (с 0), которые форматируются как деньги.
    Возвращает следующую свободную строку.
    """
    money_columns = money_columns or set()
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=col_idx, value=header)
        cell.font = Font(bold=True)

    r = start_row + 1
    for row_values in rows:
        for col_idx, value in enumerate(row_values):
            cell = ws.cell(row=r, column=col_idx + 1, value=value)
            if col_idx in money_columns and isinstance(value, (int, float)):
                cell.number_format = MONEY_FORMAT
        r += 1

    return r
