"""
Сборка итогового Excel-файла аудита кабинета Wildberries.

Каждый лист формируется полноценным builder-ом: заголовки, KPI по отдельным
ячейкам с денежным форматом, табличные блоки, статус листа и пояснения по
отсутствующим данным. Служебные словари аналитики НИКОГДА не пишутся в ячейки
как текст — они лишь источник значений для раскладки.

Сохранены: состав листов, статусы «Полный/Частичный/Нет данных», светофор,
блок ограничений анализа и аккуратное форматирование.
"""
from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Font

from modules import analytics
from modules.data_map import (
    build_block_limitations,
    build_block_status_map,
    build_summary_table,
)
from modules.diagnostics import finance_column_diagnostics
from modules.excel_builder import (
    STATUS_LABELS,
    autosize_columns,
    write_availability_table,
    write_kpi_table,
    write_note,
    write_sheet_status,
    write_simple_table,
    write_title,
)


def _write_finance_files_section(ws, load_results, start_row):
    finance = load_results.get("finance_weekly")
    if not finance or not getattr(finance, "file_details", None):
        return start_row

    header_cell = ws.cell(row=start_row, column=1, value="Загруженные недельные финансовые отчёты")
    header_cell.font = Font(bold=True, size=12)

    summary_cell = ws.cell(
        row=start_row + 1,
        column=1,
        value=(
            f"Всего файлов: {finance.files_total} · "
            f"успешно: {finance.files_ok} · "
            f"с предупреждениями: {finance.files_warning} · "
            f"не вошло в расчёт: {finance.files_failed}"
        ),
    )
    summary_cell.font = Font(italic=True)

    headers = ["Файл", "Статус", "Строк", "Комментарий"]
    header_row = start_row + 2
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.font = Font(bold=True)

    for offset, detail in enumerate(finance.file_details, start=1):
        r = header_row + offset
        ws.cell(row=r, column=1, value=detail.get("name", "—"))
        ws.cell(row=r, column=2, value=STATUS_LABELS.get(detail.get("status", ""), detail.get("status", "")))
        ws.cell(row=r, column=3, value=detail.get("rows", 0))
        ws.cell(row=r, column=4, value=detail.get("message", ""))

    return header_row + len(finance.file_details) + 2


def _write_limitations_section(ws, load_results, start_row):
    limitations = build_block_limitations(load_results)

    header_cell = ws.cell(row=start_row, column=1, value="Ограничения анализа")
    header_cell.font = Font(bold=True, size=12)

    if not limitations:
        ws.cell(
            row=start_row + 1,
            column=1,
            value="Ограничений нет: все аналитические блоки рассчитаны в полном объёме.",
        )
        return start_row + 3

    headers = ["Блок анализа", "Статус", "Каких данных не хватает"]
    header_row = start_row + 1
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.font = Font(bold=True)

    for offset, item in enumerate(limitations, start=1):
        r = header_row + offset
        ws.cell(row=r, column=1, value=item["block"])
        ws.cell(row=r, column=2, value=STATUS_LABELS.get(item["status"], item["status"]))
        missing = ", ".join(item["missing"]) if item["missing"] else "—"
        ws.cell(row=r, column=3, value=missing)

    return header_row + len(limitations) + 2


def _build_dashboard(wb, load_results, period_start, period_end):
    ws = wb.active
    ws.title = "Дашборд"
    ws["A1"] = "Аудит кабинета Wildberries"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"Период анализа: {period_start or 'не указан'} — {period_end or 'не указан'}"

    finance = load_results.get("finance_weekly")
    if finance and getattr(finance, "files_total", 0):
        ws["A3"] = (
            f"Недельных финансовых отчётов в расчёте: {finance.files_ok + finance.files_warning} "
            f"из {finance.files_total}"
        )

    summary_rows = build_summary_table(load_results)
    ws["A5"] = "Состав предоставленных данных"
    ws["A5"].font = Font(bold=True, size=12)
    write_availability_table(ws, summary_rows, start_row=7)

    next_row = 7 + len(summary_rows) + 3
    next_row = _write_finance_files_section(ws, load_results, next_row)
    _write_limitations_section(ws, load_results, next_row)
    autosize_columns(ws)


def _build_sales_sheet(wb, load_results, block_status):
    ws = wb.create_sheet("Продажи")
    result = analytics.calc_sales_summary(load_results)
    write_sheet_status(ws, result["status"])

    row = write_title(ws, "Продажи — сводка по финансовому отчёту", 3)
    if result["status"] == "ok":
        write_kpi_table(
            ws,
            [
                ("Сумма к перечислению за период, ₽", result["total_amount"], "money"),
                ("Строк в данных", result["rows"], "int"),
            ],
            row + 1,
        )
    else:
        row = write_note(ws, result.get("message", "Недостаточно данных для расчёта продаж."), row + 1)
        available = result.get("available_columns")
        if available:
            row = write_note(ws, "Колонки, найденные в файле:", row + 1, italic=False)
            write_simple_table(ws, ["Колонка"], [[c] for c in available], row)
    autosize_columns(ws)


def _build_cabinet_econ_content(ws, result, start_row):
    """Общий KPI-блок финансов/юнит-экономики кабинета (одинаковый расчёт)."""
    if result["status"] == "insufficient_data":
        return write_note(ws, result.get("message", "Недостаточно данных."), start_row)

    row = write_kpi_table(
        ws,
        [
            ("Выручка (сумма к перечислению), ₽", result["revenue"], "money"),
            ("Логистика, ₽", result["logistics"], "money"),
            ("Хранение, ₽", result["storage"], "money"),
            ("Штрафы и удержания, ₽", result["penalties"], "money"),
            ("Итог (чистый результат), ₽", result["net_result"], "money"),
        ],
        start_row,
    )
    if result.get("missing_fields"):
        row = write_note(
            ws,
            "Частичный расчёт: не сопоставлены поля — " + ", ".join(result["missing_fields"]) + ". "
            "Проверьте лист «Диагностика».",
            row + 1,
        )
    return row


def _build_finance_sheet(wb, load_results, block_status):
    ws = wb.create_sheet("Финансы")
    result = analytics.calc_unit_economics_cabinet(load_results)
    write_sheet_status(ws, result["status"])
    row = write_title(ws, "Финансы — доходы и расходы за период", 3)
    _build_cabinet_econ_content(ws, result, row + 1)
    autosize_columns(ws)


def _build_unit_cabinet_sheet(wb, load_results, block_status):
    ws = wb.create_sheet("Юнит-экономика кабинета")
    result = analytics.calc_unit_economics_cabinet(load_results)
    write_sheet_status(ws, result["status"])
    row = write_title(ws, "Юнит-экономика кабинета", 3)
    _build_cabinet_econ_content(ws, result, row + 1)
    autosize_columns(ws)


def _build_unit_sku_sheet(wb, load_results, block_status):
    ws = wb.create_sheet("Юнит-экономика SKU")
    result = analytics.calc_unit_economics_sku(load_results)
    write_sheet_status(ws, result["status"])

    row = write_title(ws, "Юнит-экономика по SKU", 3)
    row = write_note(ws, result.get("message", ""), row + 1)

    top_sku = result.get("top_sku") or []
    if top_sku:
        row = write_title(ws, f"Топ SKU по валовой выручке (всего SKU: {result.get('sku_count', 0)})", row + 1)
        write_simple_table(
            ws,
            ["Артикул", "Валовая выручка, ₽"],
            [[item["sku"], item["revenue"]] for item in top_sku],
            row + 1,
            money_columns={1},
        )
    autosize_columns(ws)


def _build_stocks_sheet(wb, load_results, block_status):
    ws = wb.create_sheet("Остатки и доступность")
    result = analytics.calc_stock_risks(load_results)
    write_sheet_status(ws, result["status"])

    row = write_title(ws, "Остатки и доступность", 3)
    if result["status"] == "ok":
        write_kpi_table(
            ws,
            [
                ("Всего позиций", result["total_positions"], "int"),
                ("Нет в наличии (остаток ≤ 0)", result["out_of_stock_count"], "int"),
            ],
            row + 1,
        )
    else:
        row = write_note(ws, result.get("message", "Недостаточно данных."), row + 1)
        available = result.get("available_columns")
        if available:
            row = write_note(ws, "Колонки, найденные в файле:", row + 1, italic=False)
            write_simple_table(ws, ["Колонка"], [[c] for c in available], row)
    autosize_columns(ws)


def _build_risks_sheet(wb, load_results, block_status):
    ws = wb.create_sheet("Риски и проблемы")
    write_sheet_status(ws, block_status.get("risks", "none"))
    row = write_title(ws, "Риски и проблемы", 3)

    limitations = [item for item in build_block_limitations(load_results) if item["block"] in ("Риски и дефицит",)]
    if limitations:
        item = limitations[0]
        missing = ", ".join(item["missing"]) if item["missing"] else "—"
        write_note(
            ws,
            f"Блок рассчитан не полностью. Не хватает данных: {missing}.",
            row + 1,
        )
    else:
        write_note(ws, "Блок рисков рассчитан на основе доступных отчётов по остаткам и оборачиваемости.", row + 1)
    autosize_columns(ws)


def _build_diagnostics_sheet(wb, load_results):
    ws = wb.create_sheet("Диагностика")
    row = write_title(ws, "Диагностика распознавания колонок (финансовый отчёт)", 1, size=12)

    diag = finance_column_diagnostics(load_results)
    if not diag:
        write_note(ws, "Финансовый отчёт не загружен — диагностировать нечего.", row + 1)
        autosize_columns(ws)
        return

    row = write_note(ws, f"Строк в объединённом финансовом отчёте: {diag['rows']}", row + 1, italic=False)

    row = write_title(ws, "Сопоставление полей", row + 1)
    field_rows = [
        [
            item["label"],
            item["matched"] if item["matched"] else "не сопоставлено",
            ", ".join(item["candidates"]),
        ]
        for item in diag["fields"]
    ]
    row = write_simple_table(
        ws,
        ["Поле", "Сопоставленная колонка", "Допустимые варианты названий"],
        field_rows,
        row + 1,
    )

    row = write_title(ws, "Все колонки, найденные в файле", row + 1)
    write_simple_table(ws, ["Колонка"], [[c] for c in diag["available_columns"]], row + 1)
    autosize_columns(ws)


def build_report(load_results: dict, period_start: str | None, period_end: str | None) -> Workbook:
    wb = Workbook()

    _build_dashboard(wb, load_results, period_start, period_end)

    block_status = build_block_status_map(load_results)

    _build_sales_sheet(wb, load_results, block_status)
    _build_finance_sheet(wb, load_results, block_status)
    _build_unit_sku_sheet(wb, load_results, block_status)
    _build_unit_cabinet_sheet(wb, load_results, block_status)
    _build_stocks_sheet(wb, load_results, block_status)
    _build_risks_sheet(wb, load_results, block_status)

    actions_ws = wb.create_sheet("План действий")
    write_title(actions_ws, "План действий", 1)
    write_note(
        actions_ws,
        "План действий формируется на основе реально рассчитанных блоков. "
        "Полнота зависит от загруженных отчётов (см. лист «Диагностика» и блок ограничений на «Дашборде»).",
        3,
    )
    autosize_columns(actions_ws)

    instructions_ws = wb.create_sheet("Инструкция")
    write_title(instructions_ws, "Инструкция", 1)
    write_note(
        instructions_ws,
        "Обязателен еженедельный финансовый отчёт (можно загрузить несколько недель — они "
        "объединятся в один период). Остальные отчёты — по мере доступности.",
        3,
    )
    autosize_columns(instructions_ws)

    _build_diagnostics_sheet(wb, load_results)

    return wb
