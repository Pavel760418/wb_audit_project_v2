"""
Генерация Excel-шаблонов для ручного заполнения пользователем.

Каждый шаблон:
- содержит лист с данными (жирная закреплённая шапка, читаемая ширина колонок);
- содержит лист «Инструкция» с понятным пояснением, как заполнять;
- включает одну строку-пример структуры (её можно удалить перед загрузкой);
- не содержит вымышленной аналитики.

Шаблоны отдаются в приложение как байты (.xlsx) для кнопки скачивания.
"""
from __future__ import annotations

import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

_HEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
_HEADER_FONT = Font(bold=True, size=11)
_TITLE_FONT = Font(bold=True, size=13)


def _write_data_sheet(ws, headers, widths, example_row=None):
    """Оформляет лист с данными: жирная шапка, закрепление, ширина колонок, пример строки."""
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    if example_row:
        for col_idx, value in enumerate(example_row, start=1):
            ws.cell(row=2, column=col_idx, value=value)

    for col_idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 30
    # Закрепляем строку заголовков.
    ws.freeze_panes = "A2"


def _write_instruction_sheet(ws, title, lines):
    ws["A1"] = title
    ws["A1"].font = _TITLE_FONT
    ws.column_dimensions["A"].width = 100
    for offset, line in enumerate(lines, start=3):
        cell = ws.cell(row=offset, column=1, value=line)
        cell.alignment = Alignment(wrap_text=True, vertical="top")


def build_cost_price_template() -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Себестоимость"
    _write_data_sheet(
        ws,
        headers=[
            "Артикул поставщика",
            "Артикул WB",
            "Наименование товара",
            "Себестоимость за единицу, ₽",
            "Валюта",
            "Комментарий",
        ],
        widths=[24, 16, 40, 26, 12, 40],
        example_row=["ПРИМЕР-001", "123456789", "Название товара", 350, "RUB", "Строку-пример можно удалить"],
    )

    instructions = wb.create_sheet("Инструкция")
    _write_instruction_sheet(
        instructions,
        "Шаблон «Себестоимость SKU»",
        [
            "Назначение: указать себестоимость каждой единицы товара для расчёта чистой прибыли по SKU.",
            "",
            "Как заполнять:",
            "1. «Артикул поставщика» — ваш внутренний артикул (обязательно). Он связывает строку с продажами.",
            "2. «Артикул WB» — артикул Wildberries (при наличии, необязательно).",
            "3. «Наименование товара» — понятное название для отчёта.",
            "4. «Себестоимость за единицу, ₽» — только число, без пробелов и знака валюты.",
            "5. «Валюта» — по умолчанию RUB.",
            "6. «Комментарий» — любые пояснения (необязательно).",
            "",
            "Важно: не переименовывайте колонки и не добавляйте лишние строки над шапкой.",
            "Строку-пример во второй строке можно удалить перед загрузкой.",
        ],
    )
    return wb


def build_extra_ads_template() -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Рекламные расходы"
    _write_data_sheet(
        ws,
        headers=[
            "Дата",
            "Артикул поставщика",
            "Кампания / Источник",
            "Тип расходов",
            "Сумма расходов, ₽",
            "Комментарий",
        ],
        widths=[16, 24, 30, 24, 22, 40],
        example_row=["2024-01-15", "ПРИМЕР-001", "Внешняя реклама", "Блогер", 5000, "Строку-пример можно удалить"],
    )

    instructions = wb.create_sheet("Инструкция")
    _write_instruction_sheet(
        instructions,
        "Шаблон «Дополнительные рекламные расходы»",
        [
            "Назначение: учесть рекламные расходы вне кабинета Wildberries (внешняя реклама, блогеры, площадки).",
            "",
            "Как заполнять:",
            "1. «Дата» — дата расхода в формате ГГГГ-ММ-ДД.",
            "2. «Артикул поставщика» — если расход относится к конкретному товару (необязательно).",
            "3. «Кампания / Источник» — где размещалась реклама.",
            "4. «Тип расходов» — например: блогер, таргет, маркетплейс-площадка.",
            "5. «Сумма расходов, ₽» — только число.",
            "6. «Комментарий» — пояснения (необязательно).",
            "",
            "Важно: не переименовывайте колонки. Строку-пример можно удалить перед загрузкой.",
        ],
    )
    return wb


def build_classifier_template() -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Классификатор"
    _write_data_sheet(
        ws,
        headers=[
            "Артикул поставщика",
            "Наименование товара",
            "Категория",
            "Бренд",
            "Группа / Коллекция",
            "Комментарий",
        ],
        widths=[24, 40, 24, 20, 26, 40],
        example_row=["ПРИМЕР-001", "Название товара", "Одежда", "Ваш бренд", "Базовая линейка", "Строку-пример можно удалить"],
    )

    instructions = wb.create_sheet("Инструкция")
    _write_instruction_sheet(
        instructions,
        "Шаблон «Классификатор SKU / Категория / Бренд»",
        [
            "Назначение: сгруппировать товары по категориям и брендам для аналитики в разрезах.",
            "",
            "Как заполнять:",
            "1. «Артикул поставщика» — ваш внутренний артикул (обязательно).",
            "2. «Наименование товара» — понятное название.",
            "3. «Категория» — товарная категория (например: одежда, обувь, аксессуары).",
            "4. «Бренд» — бренд товара.",
            "5. «Группа / Коллекция» — при необходимости (необязательно).",
            "6. «Комментарий» — пояснения (необязательно).",
            "",
            "Важно: не переименовывайте колонки. Строку-пример можно удалить перед загрузкой.",
        ],
    )
    return wb


def build_manual_adjustments_template() -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Корректировки"
    _write_data_sheet(
        ws,
        headers=[
            "Дата",
            "Артикул поставщика",
            "Тип корректировки",
            "Сумма, ₽",
            "Причина / Комментарий",
        ],
        widths=[16, 24, 28, 18, 44],
        example_row=["2024-01-20", "ПРИМЕР-001", "Возврат / компенсация", -1200, "Строку-пример можно удалить"],
    )

    instructions = wb.create_sheet("Инструкция")
    _write_instruction_sheet(
        instructions,
        "Шаблон «Ручные корректировки»",
        [
            "Назначение: внести ручные поправки к суммам, которых нет в стандартных отчётах WB.",
            "",
            "Как заполнять:",
            "1. «Дата» — дата корректировки в формате ГГГГ-ММ-ДД.",
            "2. «Артикул поставщика» — если корректировка относится к товару (необязательно).",
            "3. «Тип корректировки» — например: возврат, компенсация, ручная скидка.",
            "4. «Сумма, ₽» — число; отрицательное значение уменьшает итог, положительное увеличивает.",
            "5. «Причина / Комментарий» — обязательно поясните корректировку.",
            "",
            "Важно: не переименовывайте колонки. Строку-пример можно удалить перед загрузкой.",
        ],
    )
    return wb


# Реестр шаблонов: ключ → метаданные и функция-построитель.
TEMPLATES = {
    "cost_price": {
        "title": "Себестоимость SKU",
        "filename": "Шаблон_Себестоимость_SKU.xlsx",
        "builder": build_cost_price_template,
        "description": "Себестоимость каждой единицы товара — нужна для расчёта чистой прибыли по SKU.",
    },
    "extra_ads": {
        "title": "Дополнительные рекламные расходы",
        "filename": "Шаблон_Доп_рекламные_расходы.xlsx",
        "builder": build_extra_ads_template,
        "description": "Рекламные расходы вне кабинета Wildberries.",
    },
    "classifier": {
        "title": "Классификатор SKU / Категория / Бренд",
        "filename": "Шаблон_Классификатор_SKU.xlsx",
        "builder": build_classifier_template,
        "description": "Группировка товаров по категориям и брендам.",
    },
    "manual_adjustments": {
        "title": "Ручные корректировки",
        "filename": "Шаблон_Ручные_корректировки.xlsx",
        "builder": build_manual_adjustments_template,
        "description": "Ручные поправки к суммам, которых нет в стандартных отчётах.",
    },
}


def build_template_bytes(template_key: str) -> bytes:
    """Возвращает готовый .xlsx-шаблон в виде байтов для скачивания."""
    meta = TEMPLATES[template_key]
    workbook = meta["builder"]()
    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
